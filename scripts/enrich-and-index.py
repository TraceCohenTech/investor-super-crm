#!/usr/bin/env python3
"""
NYVP Investor CRM — Enrichment & Unified Search Index Builder (v2)

Primary source: NYVP_Investor_CRM.xlsx (master Excel with curated tags)
Fallback: existing JSON data files for records not in Excel
Outputs: search-index.json, search-meta.json, dedup-candidates.json
"""

import json
import os
import re
import hashlib
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from difflib import SequenceMatcher

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(["pip3", "install", "openpyxl"])
    import openpyxl

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "NYVP_Investor_CRM.xlsx")
TODAY = datetime.now()

# ─── Sector normalization ───
SECTOR_NORMALIZE = {
    "ai": "AI/ML", "ml": "AI/ML", "machine learning": "AI/ML", "artificial intelligence": "AI/ML",
    "llm": "AI/ML", "generative ai": "AI/ML", "deep learning": "AI/ML", "applied ai": "Applied AI",
    "deep tech": "Deep Tech", "defense tech": "Defense Tech",
    "saas": "B2B SaaS", "enterprise": "B2B SaaS", "enterprise saas": "B2B SaaS",
    "fintech": "Fintech", "insurtech": "Fintech",
    "health": "Healthcare", "biotech": "Healthcare", "medtech": "Healthcare",
    "climate": "Climate", "sustainability": "Climate", "cleantech": "Climate",
    "crypto": "Crypto/Web3", "web3": "Crypto/Web3", "blockchain": "Crypto/Web3",
    "defense": "Defense/Gov", "military": "Defense/Gov", "govtech": "Defense/Gov",
    "consumer": "Consumer", "d2c": "Consumer", "e-commerce": "Consumer",
    "proptech": "PropTech", "real estate": "PropTech", "construction tech": "PropTech",
    "edtech": "EdTech", "education": "EdTech",
    "b2b saas": "B2B SaaS",
}

# Known short sector labels to keep as-is
KNOWN_SECTORS = {
    "AI/ML", "Applied AI", "Deep Tech", "Defense Tech", "B2B SaaS", "Fintech",
    "Healthcare", "Climate", "Crypto/Web3", "Defense/Gov", "Consumer", "PropTech",
    "EdTech", "Media/Content", "Generalist", "Multi-Stage", "Credit/PE",
    "Tech", "Enterprise SaaS",
}

# ─── Role type inference ───
ROLE_MAP = [
    (["general partner", "managing partner", " gp ", "gp,", "gp/"], "GP/Partner"),
    (["partner"], "Partner"),
    (["principal"], "Principal"),
    (["associate", "analyst"], "Associate/Analyst"),
    (["founder", "co-founder", "cofounder", "ceo", "chief executive"], "Founder/CEO"),
    (["cto", "cfo", "coo", "chief"], "C-Suite"),
    (["vp ", "vice president", "svp", "evp"], "VP"),
    (["director"], "Director"),
    (["managing director", " md "], "Managing Director"),
]

# ─── Nickname variants for fuzzy dedup ───
NICKNAMES = {
    "robert": ["rob", "bob", "bobby"], "william": ["will", "bill", "billy"],
    "james": ["jim", "jimmy"], "richard": ["rick", "dick"],
    "michael": ["mike", "mick"], "joseph": ["joe"],
    "thomas": ["tom", "tommy"], "christopher": ["chris"],
    "matthew": ["matt"], "daniel": ["dan", "danny"],
    "nicholas": ["nick"], "benjamin": ["ben"],
    "jonathan": ["jon", "john"], "anthony": ["tony"],
    "elizabeth": ["liz", "beth"], "jennifer": ["jen"],
    "katherine": ["kate", "kathy", "katie"], "alexander": ["alex"],
    "samuel": ["sam"], "nathaniel": ["nate", "nathan"],
    "joshua": ["josh"], "andrew": ["andy", "drew"],
    "timothy": ["tim"], "stephen": ["steve"],
    "david": ["dave"], "edward": ["ed", "ted"],
    "gregory": ["greg"], "lawrence": ["larry"],
    "phillip": ["phil"], "jp": ["j.p."], "j.p.": ["jp"],
}


def load_json(filename):
    path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def clean_name(name):
    if not name:
        return ""
    name = re.sub(r",?\s*(Ph\.?D\.?|MD|MBA|CFA|CPA|Esq\.?|Jr\.?|Sr\.?|III|II|IV)\s*$", "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s+", " ", name).strip()
    if "@" in name:
        return ""
    return name


def normalize_sector(raw):
    """Normalize a sector string to a short label."""
    if not raw or raw == "None" or raw == "Unknown":
        return "Generalist"
    raw = raw.strip()
    # Already a known short label
    if raw in KNOWN_SECTORS:
        return raw
    # Try keyword matching on long descriptions
    lower = raw.lower()
    for keyword, label in sorted(SECTOR_NORMALIZE.items(), key=lambda x: -len(x[0])):
        if keyword in lower:
            return label
    # If it's short enough, keep it
    if len(raw) <= 25:
        return raw
    return "Generalist"


def infer_role_type(title):
    if not title:
        return ""
    t = f" {title.lower()} "
    for keywords, role in ROLE_MAP:
        if any(k in t for k in keywords):
            return role
    return ""


def parse_check_size(raw):
    if not raw:
        return ""
    raw = raw.lower().replace(",", "").replace(" ", "")
    amounts = re.findall(r'\$?([\d.]+)\s*([kmb])?', raw)
    if not amounts:
        return ""
    parsed = []
    for val, unit in amounts:
        try:
            n = float(val)
            if unit == "k":
                n *= 1000
            elif unit == "m":
                n *= 1_000_000
            elif unit == "b":
                n *= 1_000_000_000
            elif n < 100:
                n *= 1_000_000
            parsed.append(n)
        except ValueError:
            continue
    if not parsed:
        return ""
    max_val = max(parsed)
    if max_val < 100_000:
        return "<$100K"
    if max_val < 500_000:
        return "$100K-$500K"
    if max_val < 1_000_000:
        return "$500K-$1M"
    if max_val < 5_000_000:
        return "$1M-$5M"
    return "$5M+"


def compute_strength(email_count, wa_groups_count, sources_count, in_crm):
    if email_count >= 5 or sources_count >= 4:
        return "Strong"
    if email_count >= 2 or wa_groups_count >= 1 or sources_count >= 2:
        return "Medium"
    if email_count >= 1 or in_crm:
        return "Weak"
    return "None"


def compute_quality_grade(entry):
    """Compute completeness score and letter grade."""
    score = 0
    if entry.get("n"):                                      score += 10
    if entry.get("e"):                                      score += 20
    if entry.get("c"):                                      score += 10
    if entry.get("t"):                                      score += 10
    if entry.get("li"):                                     score += 10
    if entry.get("rg"):                                     score += 10
    if entry.get("it") and entry["it"] != "Other":          score += 10
    if entry.get("fs") and entry["fs"] != "Unknown":        score += 5
    if entry.get("sc") and entry["sc"] != ["Generalist"]:   score += 5
    if entry.get("cs"):                                     score += 5
    if entry.get("rs") and entry["rs"] != "None":           score += 5
    if score >= 80:   return "A"
    elif score >= 60: return "B"
    elif score >= 40: return "C"
    elif score >= 20: return "D"
    else:             return "F"


def compute_staleness(status, last_contact_str, email_count, rs):
    """Compute staleness from last contact date."""
    if not last_contact_str or last_contact_str == "None":
        if status in ("Active", "Warm"):
            return "at-risk"
        return "unknown"

    # Parse last contact date
    lc_date = None
    for fmt in ["%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%b %d, %Y", "%B %d, %Y"]:
        try:
            lc_date = datetime.strptime(last_contact_str[:10], fmt[:min(len(fmt), 10)])
            break
        except (ValueError, IndexError):
            continue
    if not lc_date:
        # Try partial dates like "Sun, 01 Ma"
        try:
            lc_date = datetime.strptime(last_contact_str[:10], "%Y-%m-%d")
        except (ValueError, IndexError):
            if status in ("Active", "Warm"):
                return "at-risk"
            return "unknown"

    days_since = (TODAY - lc_date).days

    if status in ("Active", "Warm"):
        if days_since > 90:
            return "stale"
        elif days_since > 30:
            return "at-risk"
        else:
            return "healthy"
    elif email_count >= 2:
        if days_since > 180:
            return "stale"
        elif days_since > 90:
            return "at-risk"
        else:
            return "healthy"
    elif rs in ("Strong", "Medium"):
        return "healthy"
    return "unknown"


def normalize_company_for_dedup(company):
    if not company:
        return ""
    c = company.lower().strip()
    for suffix in [" inc", " inc.", " llc", " corp", " corp.", " co.", " ltd", " ltd.",
                   " lp", " l.p.", " gp", " ventures", " capital", " partners",
                   ", inc", ", llc", ", corp", " management", " advisors", " group"]:
        c = c.replace(suffix, "")
    c = re.sub(r'[.\-\'\"]', '', c)
    c = re.sub(r'\s+', ' ', c).strip()
    return c


def names_fuzzy_match(name1, name2, threshold=0.85):
    if not name1 or not name2:
        return False, 0.0
    n1 = name1.lower().strip()
    n2 = name2.lower().strip()
    if n1 == n2:
        return True, 1.0
    ratio = SequenceMatcher(None, n1, n2).ratio()
    if ratio >= threshold:
        return True, ratio
    # Check nickname variants
    parts1 = n1.split()
    parts2 = n2.split()
    if len(parts1) >= 2 and len(parts2) >= 2 and parts1[-1] == parts2[-1]:
        first1, first2 = parts1[0], parts2[0]
        for full, nicks in NICKNAMES.items():
            all_forms = [full] + nicks
            if first1 in all_forms and first2 in all_forms:
                return True, 0.90
    return False, ratio


# ═══════════════════════════════════════════════════
# LOAD PRIMARY SOURCE: Excel
# ═══════════════════════════════════════════════════
print("Loading Excel master file...")

# Map Excel sheet names to source identifiers
SHEET_SOURCE_MAP = {
    "🔥 Follow-Up Now": "follow-up",
    "🏦 Investors & Funds": "investors",
    "👼 Angels": "angels",
    "✅ LinkedIn Verified": "linkedin-verified",
    "💬 WhatsApp Contacts": "whatsapp",
    "🇮🇱 Israel": "israel",
    "🗽 NYC Investors": "nyc",
    "🌴 South Florida": "south-florida",
    "💼 LinkedIn Network": "linkedin",
    "💰 LP Pipeline": "lp-pipeline",
    "🎯 Re-Engage": "re-engage",
    "🚀 Deal Flow": "dealflow",
    "⚖️ Legal & Services": "legal",
    "📋 Needs Review": "needs-review",
}

raw_records = []

if os.path.exists(EXCEL_PATH):
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    for sheet_name, source_id in SHEET_SOURCE_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        headers = [str(h or "") for h in rows[0]]
        col = {h: i for i, h in enumerate(headers)}

        for row in rows[1:]:
            def cell(h):
                idx = col.get(h)
                if idx is None or idx >= len(row):
                    return ""
                v = row[idx]
                if v is None:
                    return ""
                return str(v).strip() if isinstance(v, str) else v

            name = clean_name(str(cell("Name") or ""))
            email = str(cell("Email") or "").strip().lower()
            if not name and not email:
                continue

            # Parse WA Groups (pipe-separated)
            wa_raw = str(cell("WA Groups") or "")
            wa_groups = [g.strip() for g in wa_raw.split("|") if g.strip() and g.strip() != "None"]

            # Parse tags
            tags_raw = str(cell("Tags") or "")
            tags = [t.strip() for t in tags_raw.split() if t.strip() and t.strip() != "None"]

            # Email count
            ec_raw = cell("Emails")
            try:
                email_count = int(ec_raw) if ec_raw and str(ec_raw) != "None" else 0
            except (ValueError, TypeError):
                email_count = 0

            # Last contact
            lc_raw = cell("Last Contact")
            lc_str = ""
            if lc_raw and str(lc_raw) != "None":
                if hasattr(lc_raw, "strftime"):
                    lc_str = lc_raw.strftime("%Y-%m-%d")
                else:
                    lc_str = str(lc_raw)[:10]

            # Follow-up
            fu_raw = cell("Follow-Up")
            fu_str = ""
            if fu_raw and str(fu_raw) != "None":
                if hasattr(fu_raw, "strftime"):
                    fu_str = fu_raw.strftime("%Y-%m-%d")
                else:
                    fu_str = str(fu_raw)[:10]

            # Priority
            pr_raw = str(cell("Priority") or "")
            priority = pr_raw if pr_raw in ("High", "Medium", "Low") else ""

            # LinkedIn verified
            li_verified = str(cell("LI ✓") or "")
            lv = "verified" if li_verified == "✓" else ("unverified" if li_verified == "~" else "")

            # Israel flag
            il_raw = str(cell("Israel") or "")
            israel = bool(il_raw and il_raw != "None")

            # Category from Excel (much richer than our inference)
            category = str(cell("Category") or "")
            if category == "None":
                category = ""

            # Status
            status = str(cell("Status") or "")
            if status == "None":
                status = ""

            # Geo
            geo = str(cell("Geo") or "")
            if geo == "None":
                geo = ""

            # Fund Stage
            fund_stage = str(cell("Fund Stage") or "")
            if fund_stage in ("None", "Unknown"):
                fund_stage = ""

            # Sector
            sector_raw = str(cell("Sector") or "")
            if sector_raw in ("None", "Unknown"):
                sector_raw = ""

            # Notes
            notes = str(cell("Notes") or "")
            if notes == "None":
                notes = ""

            # Source string
            source_str = str(cell("Source") or "")
            if source_str == "None":
                source_str = ""

            # HS Status
            hs_status = str(cell("HS Status") or "")
            if hs_status == "None":
                hs_status = ""

            rec = {
                "_source": "excel",
                "_sheet": source_id,
                "name": name,
                "email": email if "@" in email else "",
                "company": str(cell("Company") or "").strip(),
                "title": str(cell("Title") or "").strip() if str(cell("Title") or "") != "None" else "",
                "category": category,
                "status": status,
                "last_contact": lc_str,
                "email_count": email_count,
                "follow_up": fu_str,
                "priority": priority,
                "linkedin": str(cell("LinkedIn") or "").strip() if str(cell("LinkedIn") or "") != "None" else "",
                "li_verified": lv,
                "wa_groups": wa_groups,
                "fund_stage": fund_stage,
                "sector_raw": sector_raw,
                "geo": geo,
                "israel": israel,
                "hs_status": hs_status,
                "tags": tags,
                "notes": notes,
                "source_str": source_str,
            }
            raw_records.append(rec)

    wb.close()
    print(f"  Loaded {len(raw_records)} rows from Excel ({len(SHEET_SOURCE_MAP)} sheets)")
else:
    print("  Excel file not found, falling back to JSON sources only")

# ═══════════════════════════════════════════════════
# DEDUP: By email first, then name+company
# ═══════════════════════════════════════════════════
print("Deduplicating...")

# Sheet priority: determines which sheet's data wins during merge
SHEET_PRIORITY = [
    "follow-up", "investors", "nyc", "south-florida", "israel", "lp-pipeline",
    "angels", "whatsapp", "linkedin-verified", "linkedin", "re-engage",
    "dealflow", "legal", "needs-review"
]

email_groups = defaultdict(list)
no_email = []

for rec in raw_records:
    email = rec.get("email", "")
    if email and "@" in email:
        email_groups[email].append(rec)
    else:
        no_email.append(rec)


def sheet_priority_key(rec):
    sheet = rec.get("_sheet", "")
    return SHEET_PRIORITY.index(sheet) if sheet in SHEET_PRIORITY else 99


def merge_excel_records(records):
    """Merge multiple Excel records for the same person."""
    records.sort(key=sheet_priority_key)
    merged = {}
    all_sheets = set()
    all_tags = set()
    max_email_count = 0
    all_wa_groups = set()
    all_notes = []

    for rec in records:
        all_sheets.add(rec["_sheet"])
        all_tags.update(rec.get("tags", []))
        max_email_count = max(max_email_count, rec.get("email_count", 0))
        all_wa_groups.update(rec.get("wa_groups", []))
        if rec.get("notes"):
            all_notes.append(rec["notes"])

        # Take first non-empty value for scalar fields
        for field in ["name", "email", "company", "title", "category", "status",
                       "last_contact", "follow_up", "priority", "linkedin",
                       "li_verified", "fund_stage", "sector_raw", "geo",
                       "israel", "hs_status", "source_str"]:
            val = rec.get(field, "")
            if val and val != "Unknown" and field not in merged:
                merged[field] = val

    merged["_sheets"] = sorted(all_sheets)
    merged["tags"] = sorted(all_tags)
    merged["email_count"] = max_email_count
    merged["wa_groups"] = sorted(all_wa_groups)
    merged["notes"] = " | ".join(all_notes[:3]) if all_notes else ""
    return merged


# Merge email groups
unified = []
for email, recs in email_groups.items():
    m = merge_excel_records(recs)
    m["email"] = email
    m["name"] = clean_name(m.get("name", ""))
    unified.append(m)

# Dedup no-email records by name+company
name_groups = defaultdict(list)
for rec in no_email:
    name = clean_name(rec.get("name", "")).lower()
    company = (rec.get("company") or "").strip().lower()
    if name:
        key = f"{name}||{company}"
        name_groups[key].append(rec)

for key, recs in name_groups.items():
    m = merge_excel_records(recs)
    m["name"] = clean_name(m.get("name", ""))
    # Check if already in unified (from email pass)
    name_lower = m["name"].lower()
    company_lower = (m.get("company") or "").lower()
    duplicate = False
    for u in unified:
        if u.get("name", "").lower() == name_lower and (u.get("company") or "").lower() == company_lower:
            # Merge into existing
            for field in ["title", "geo", "fund_stage", "sector_raw", "status",
                          "priority", "linkedin", "category", "last_contact", "source_str"]:
                if not u.get(field) and m.get(field):
                    u[field] = m[field]
            u["_sheets"] = sorted(set(u.get("_sheets", [])) | set(m.get("_sheets", [])))
            u["tags"] = sorted(set(u.get("tags", [])) | set(m.get("tags", [])))
            u["email_count"] = max(u.get("email_count", 0), m.get("email_count", 0))
            u["wa_groups"] = sorted(set(u.get("wa_groups", [])) | set(m.get("wa_groups", [])))
            duplicate = True
            break
    if not duplicate:
        unified.append(m)

print(f"  After dedup: {len(unified)} unique records")

# ═══════════════════════════════════════════════════
# ENRICH: Build final index entries
# ═══════════════════════════════════════════════════
print("Enriching records...")

# Map sheet names to display-friendly source labels
SHEET_TO_SOURCE = {
    "follow-up": "follow-up", "investors": "investors", "nyc": "nyc",
    "south-florida": "south-florida", "israel": "israel", "lp-pipeline": "lp-pipeline",
    "angels": "angels", "whatsapp": "whatsapp", "linkedin-verified": "linkedin-verified",
    "linkedin": "linkedin", "re-engage": "re-engage", "dealflow": "dealflow",
    "legal": "legal", "needs-review": "needs-review",
}

final = []
for rec in unified:
    name = rec.get("name", "")
    if not name:
        continue

    # Stable ID
    id_seed = rec.get("email") or f"{name}|{rec.get('company', '')}"
    rid = hashlib.md5(id_seed.lower().encode()).hexdigest()[:12]

    # Build source list from sheets
    sheets = rec.get("_sheets", [])
    src = sorted(set(SHEET_TO_SOURCE.get(s, s) for s in sheets))
    # Also parse source_str for additional sources
    source_str = rec.get("source_str", "")
    if source_str:
        if "hubspot" in source_str.lower() or "HubSpot" in source_str:
            src = sorted(set(src) | {"hubspot"})
        if "gmail" in source_str.lower():
            src = sorted(set(src) | {"gmail"})
        if "calendar" in source_str.lower():
            src = sorted(set(src) | {"calendar"})
        if "linkedin" in source_str.lower() and "linkedin" not in src and "linkedin-verified" not in src:
            src = sorted(set(src) | {"linkedin"})

    # Category (use Excel's rich categories)
    category = rec.get("category", "Other") or "Other"

    # Fund stage - normalize
    fund_stage = rec.get("fund_stage", "")
    # Check if fund stage contains a dollar range
    check_size = ""
    if fund_stage and "$" in fund_stage:
        check_size = parse_check_size(fund_stage)
        fund_stage = ""  # dollar ranges go to check_size, not fund_stage
    elif fund_stage:
        # Normalize stage names
        fs_lower = fund_stage.lower()
        if "angel" in fs_lower: fund_stage = "Angel"
        elif "pre-seed" in fs_lower and "seed" in fs_lower and "series" not in fs_lower: fund_stage = "Pre-Seed/Seed"
        elif "pre-seed" in fs_lower and "series a" in fs_lower: fund_stage = "Pre-Seed - Series A"
        elif "pre-seed" in fs_lower: fund_stage = "Pre-Seed"
        elif "seed" in fs_lower and "series b" in fs_lower: fund_stage = "Seed - Series B+"
        elif "seed/a" in fs_lower or ("seed" in fs_lower and "series a" in fs_lower): fund_stage = "Seed/A"
        elif "seed" in fs_lower: fund_stage = "Seed"
        elif "series a/b" in fs_lower or "series a" in fs_lower: fund_stage = "Series A/B"
        elif "series b" in fs_lower: fund_stage = "Series B+"
        elif "multi" in fs_lower: fund_stage = "Multi-Stage"
        elif "growth" in fs_lower: fund_stage = "Growth"
        elif "pe" in fs_lower or "private equity" in fs_lower: fund_stage = "PE"
        elif "debt" in fs_lower or "credit" in fs_lower: fund_stage = "Debt/Credit"

    # Sector normalization
    sector_raw = rec.get("sector_raw", "")
    sector = normalize_sector(sector_raw)
    sectors = [sector] if sector != "Generalist" else []
    # Also check tags for sector hints
    tags = rec.get("tags", [])
    tag_sector_map = {
        "ai": "AI/ML", "deep-tech": "Deep Tech", "applied-ai-group": "Applied AI",
        "defense": "Defense/Gov", "defense-tech-group": "Defense Tech",
        "fintech": "Fintech", "crypto": "Crypto/Web3", "saas": "B2B SaaS",
        "consumer": "Consumer", "climate": "Climate", "healthcare": "Healthcare",
        "infra": "B2B SaaS",
    }
    for tag in tags:
        if tag in tag_sector_map and tag_sector_map[tag] not in sectors:
            sectors.append(tag_sector_map[tag])
    if not sectors:
        sectors = ["Generalist"]

    # Region
    geo = rec.get("geo", "")
    if not geo:
        # Try to infer from tags
        for tag in tags:
            if tag == "nyc": geo = "NYC Metro"; break
            elif tag == "bay-area": geo = "Bay Area"; break
            elif tag == "boston": geo = "Boston"; break
            elif tag == "la": geo = "LA / SoCal"; break
            elif tag == "south-florida": geo = "South Florida"; break
            elif tag == "israel": geo = "Israel"; break

    # Relationship strength
    email_count = rec.get("email_count", 0)
    wa_count = len(rec.get("wa_groups", []))
    in_crm = len(src) >= 2 or "hubspot" in src or "investors" in src
    rs = compute_strength(email_count, wa_count, len(src), in_crm)

    # Status normalization
    status = rec.get("status", "")
    if status not in ("Active", "Warm", "Contacted", "New"):
        if status:
            s = status.lower()
            if "active" in s: status = "Active"
            elif "warm" in s: status = "Warm"
            elif "contact" in s: status = "Contacted"
            elif "new" in s: status = "New"
            else: status = ""
        else:
            status = ""

    # CRM status
    crm = "In CRM" if in_crm else "New"

    # LinkedIn URL
    li = rec.get("linkedin", "")
    if li and not li.startswith("http"):
        li = ""

    # ─── Smart Auto-Tags ───
    auto_tags = set(tags)
    title_lower = (rec.get("title") or "").lower()
    company_lower = (rec.get("company") or "").lower()

    # Accelerator / notable org networks
    yc_keywords = ["y combinator", "ycombinator", " yc ", "yc "]
    if any(k in company_lower or k in title_lower for k in yc_keywords):
        auto_tags.add("yc-network")
    if any(k in company_lower for k in ["500 startups", "500 global", "500.co"]):
        auto_tags.add("500-startups")
    if any(k in company_lower for k in ["techstars"]):
        auto_tags.add("techstars")

    # Big tech alumni
    big_tech = ["google", "meta", "facebook", "amazon", "apple", "microsoft",
                "netflix", "stripe", "uber", "airbnb", "tesla", "spacex"]
    if any(k in company_lower for k in big_tech):
        auto_tags.add("big-tech")

    # Board members
    if any(k in title_lower for k in ["board member", "board director", "board of directors",
                                        "advisory board", "board advisor"]):
        auto_tags.add("board-member")

    # Operator-investors
    if ("ceo" in title_lower or "founder" in title_lower or "cto" in title_lower) and \
       ("angel" in title_lower or "investor" in title_lower or "advisor" in title_lower):
        auto_tags.add("operator-investor")

    # Reachability
    if rec.get("email") and li:
        auto_tags.add("reachable")
    elif rec.get("email"):
        auto_tags.add("email-only")
    elif li:
        auto_tags.add("linkedin-only")

    # Engagement-based tags
    if email_count >= 10:
        auto_tags.add("high-engagement")
    if len(rec.get("wa_groups", [])) >= 2:
        auto_tags.add("multi-wa")

    # Recent activity
    lc_str = rec.get("last_contact", "")
    if lc_str and lc_str != "None":
        try:
            lc_dt = datetime.strptime(lc_str[:10], "%Y-%m-%d")
            days_ago = (TODAY - lc_dt).days
            if days_ago <= 7:
                auto_tags.add("recently-active")
            elif days_ago <= 30:
                auto_tags.add("active-30d")
        except (ValueError, IndexError):
            pass

    # Multi-source (appears in 3+ sources)
    if len(src) >= 3:
        auto_tags.add("multi-source")

    tags = sorted(auto_tags)

    entry = {
        "id": rid,
        "n": name,
        "e": rec.get("email", ""),
        "c": rec.get("company", ""),
        "t": rec.get("title", ""),
        "li": li,
        "lv": rec.get("li_verified", ""),
        "src": src,
        "so": rec.get("source_str", ""),
        "st": status,
        "rg": geo,
        "fs": fund_stage,
        "sc": sectors,
        "pr": rec.get("priority", ""),
        "crm": crm,
        "it": category,
        "rt": infer_role_type(rec.get("title", "")),
        "cs": check_size,
        "rs": rs,
        "wg": rec.get("wa_groups", []),
        "tg": tags,
        "ec": email_count,
        "lc": rec.get("last_contact", ""),
        "fu": rec.get("follow_up", ""),
        "nt": rec.get("notes", ""),
        "il": rec.get("israel", False),
    }

    # Compute quality grade
    entry["q"] = compute_quality_grade(entry)

    # Compute staleness
    entry["sl"] = compute_staleness(status, rec.get("last_contact", ""), email_count, rs)

    final.append(entry)

print(f"  Enriched {len(final)} records")

# ═══════════════════════════════════════════════════
# FUZZY DEDUP: Find candidates (don't auto-merge)
# ═══════════════════════════════════════════════════
print("Finding fuzzy duplicate candidates...")

company_groups = defaultdict(list)
for i, rec in enumerate(final):
    nc = normalize_company_for_dedup(rec["c"])
    if nc and len(nc) >= 2:
        company_groups[nc].append(i)

dedup_candidates = []
seen_pairs = set()

for company_key, indices in company_groups.items():
    if len(indices) < 2 or len(indices) > 50:  # skip very large groups
        continue
    for i in range(len(indices)):
        for j in range(i + 1, len(indices)):
            r1 = final[indices[i]]
            r2 = final[indices[j]]
            if r1["id"] == r2["id"]:
                continue
            # Skip if same email (already deduped)
            if r1["e"] and r1["e"] == r2["e"]:
                continue
            pair_key = tuple(sorted([r1["id"], r2["id"]]))
            if pair_key in seen_pairs:
                continue

            match, similarity = names_fuzzy_match(r1["n"], r2["n"])
            if match:
                seen_pairs.add(pair_key)
                dedup_candidates.append({
                    "id1": r1["id"], "name1": r1["n"], "email1": r1["e"],
                    "company1": r1["c"], "title1": r1["t"], "sources1": r1["src"],
                    "grade1": r1["q"],
                    "id2": r2["id"], "name2": r2["n"], "email2": r2["e"],
                    "company2": r2["c"], "title2": r2["t"], "sources2": r2["src"],
                    "grade2": r2["q"],
                    "similarity": round(similarity, 2),
                })

# Sort by similarity desc
dedup_candidates.sort(key=lambda x: -x["similarity"])
print(f"  Found {len(dedup_candidates)} fuzzy duplicate candidates")

# ═══════════════════════════════════════════════════
# COMPUTE FACETS
# ═══════════════════════════════════════════════════
print("Computing facets...")
facets = {}

facet_dims = {
    "investorType": "it",
    "region": "rg",
    "fundStage": "fs",
    "status": "st",
    "crmStatus": "crm",
    "relationshipStrength": "rs",
    "roleType": "rt",
    "checkSize": "cs",
    "qualityGrade": "q",
    "staleness": "sl",
}

for facet_name, field in facet_dims.items():
    counts = Counter(r[field] for r in final if r[field] and r[field] != "unknown")
    facets[facet_name] = dict(counts.most_common(30))

# Sectors (arrays)
sector_counter = Counter()
for r in final:
    for s in r["sc"]:
        sector_counter[s] += 1
facets["sector"] = dict(sector_counter.most_common(25))

# Sources (arrays)
source_counter = Counter()
for r in final:
    for s in r["src"]:
        source_counter[s] += 1
facets["sources"] = dict(source_counter.most_common(25))

# Tags (arrays) — top 30
tag_counter = Counter()
for r in final:
    for t in r["tg"]:
        tag_counter[t] += 1
facets["tags"] = dict(tag_counter.most_common(40))

# Data quality summary
grade_dist = Counter(r["q"] for r in final)
total = len(final)
quality_summary = {
    "gradeDistribution": {g: grade_dist.get(g, 0) for g in ["A", "B", "C", "D", "F"]},
    "overallScore": round(sum(1 for r in final if r["q"] in ("A", "B")) / total * 100, 1) if total else 0,
    "topGaps": sorted([
        {"field": "Email", "missing": sum(1 for r in final if not r["e"]), "pct": round(sum(1 for r in final if not r["e"]) / total * 100, 1)},
        {"field": "Region", "missing": sum(1 for r in final if not r["rg"]), "pct": round(sum(1 for r in final if not r["rg"]) / total * 100, 1)},
        {"field": "Fund Stage", "missing": sum(1 for r in final if not r["fs"]), "pct": round(sum(1 for r in final if not r["fs"]) / total * 100, 1)},
        {"field": "Title", "missing": sum(1 for r in final if not r["t"]), "pct": round(sum(1 for r in final if not r["t"]) / total * 100, 1)},
        {"field": "LinkedIn", "missing": sum(1 for r in final if not r["li"]), "pct": round(sum(1 for r in final if not r["li"]) / total * 100, 1)},
        {"field": "Company", "missing": sum(1 for r in final if not r["c"]), "pct": round(sum(1 for r in final if not r["c"]) / total * 100, 1)},
        {"field": "Sector", "missing": sum(1 for r in final if r["sc"] == ["Generalist"]), "pct": round(sum(1 for r in final if r["sc"] == ["Generalist"]) / total * 100, 1)},
    ], key=lambda x: -x["pct"]),
}

# Staleness summary
staleness_summary = {
    "stale": sum(1 for r in final if r["sl"] == "stale"),
    "atRisk": sum(1 for r in final if r["sl"] == "at-risk"),
    "healthy": sum(1 for r in final if r["sl"] == "healthy"),
    "unknown": sum(1 for r in final if r["sl"] == "unknown"),
}

# ═══════════════════════════════════════════════════
# WRITE OUTPUT
# ═══════════════════════════════════════════════════
print("Writing output files...")

index_path = os.path.join(DATA_DIR, "search-index.json")
meta_path = os.path.join(DATA_DIR, "search-meta.json")
dedup_path = os.path.join(DATA_DIR, "dedup-candidates.json")

# Strip empty strings and false booleans to save space
def compact(entry):
    return {k: v for k, v in entry.items() if v or v == 0 or k == "id"}

compact_records = [compact(r) for r in final]

with open(index_path, "w", encoding="utf-8") as f:
    json.dump({"records": compact_records, "facets": facets}, f, separators=(",", ":"))

with open(meta_path, "w", encoding="utf-8") as f:
    json.dump({
        "total": len(final),
        "facets": facets,
        "dataQuality": quality_summary,
        "staleness": staleness_summary,
        "dedupCandidates": len(dedup_candidates),
    }, f, indent=2)

with open(dedup_path, "w", encoding="utf-8") as f:
    json.dump(dedup_candidates, f, indent=2)

# ═══════════════════════════════════════════════════
# SUMMARY
# ═══════════════════════════════════════════════════
size_mb = os.path.getsize(index_path) / (1024 * 1024)
print(f"\n{'='*50}")
print(f"Output: {index_path} ({size_mb:.1f} MB, {len(final)} records)")
print(f"Dedup candidates: {len(dedup_candidates)}")
print(f"\n=== Quality Grades ===")
for g in ["A", "B", "C", "D", "F"]:
    print(f"  {g}: {grade_dist.get(g, 0):,}")
print(f"  Overall score: {quality_summary['overallScore']}% (A+B)")
print(f"\n=== Staleness ===")
for k, v in staleness_summary.items():
    print(f"  {k}: {v:,}")
print(f"\n=== Top Categories ===")
for k, v in Counter(r["it"] for r in final).most_common(10):
    print(f"  {k}: {v:,}")
print(f"\n=== Top Regions ===")
for k, v in Counter(r["rg"] for r in final if r["rg"]).most_common(10):
    print(f"  {k}: {v:,}")
print(f"\n=== Top Tags ===")
for k, v in tag_counter.most_common(15):
    print(f"  {k}: {v:,}")
